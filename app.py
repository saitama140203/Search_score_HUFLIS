#!/usr/bin/env python3
"""Streamlit app quản lý điểm ĐHNN - không dùng pandas."""
import streamlit as st
from pathlib import Path
from openpyxl import load_workbook
import csv
from collections import Counter
import json

# Cấu hình trang
st.set_page_config(
    page_title="Hệ thống quản lý điểm - ĐHNN (No Pandas)",
    page_icon="📊",
    layout="wide"
)

class DataProcessor:
    def __init__(self, base_path="data_diem_dhnn"):
        self.base_path = Path(base_path)
        self.processing_path = self.base_path / "processing"
        self.processing_path.mkdir(exist_ok=True)
    
    def load_data_as_dict(self):
        """Đọc dữ liệu từ Excel thành dict."""
        excel_path = self.processing_path / "output_direct.xlsx"
        
        if not excel_path.exists():
            return None, "Không tìm thấy file output_direct.xlsx"
        
        try:
            st.info("🔧 Đang đọc dữ liệu bằng openpyxl...")
            
            # Tạo CSV tạm
            csv_path = self.processing_path / "temp_no_pandas.csv"
            
            wb = load_workbook(str(excel_path), read_only=True, data_only=True)
            ws = wb.active
            
            # Ghi CSV
            with open(str(csv_path), 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    clean_row = [str(cell) if cell is not None else "" for cell in row]
                    writer.writerow(clean_row)
            
            wb.close()
            
            # Đọc CSV thành dict
            data = []
            with open(str(csv_path), 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    # Lọc dòng có mã SV hợp lệ
                    if len(row.get('Mã SV', '')) > 5:
                        data.append(dict(row))
            
            # Xóa file tạm
            csv_path.unlink()
            
            st.success(f"✅ Đã đọc {len(data):,} bản ghi!")
            return data, None
            
        except Exception as e:
            return None, f"Lỗi: {str(e)}"
    
    def analyze_data(self, data):
        """Phân tích dữ liệu."""
        if not data:
            return {}
        
        stats = {
            'total_records': len(data),
            'by_semester': Counter(),
            'by_khoa': Counter(),
            'by_subject': Counter(),
            'scores': []
        }
        
        for record in data:
            # Thống kê theo học kỳ
            if 'Học kỳ' in record:
                stats['by_semester'][record['Học kỳ']] += 1
            
            # Thống kê theo khóa
            if 'Khóa' in record:
                stats['by_khoa'][record['Khóa']] += 1
            
            # Thống kê theo môn
            if 'Môn học' in record:
                stats['by_subject'][record['Môn học']] += 1
            
            # Thu thập điểm
            if 'Điểm TBTL' in record:
                try:
                    score = float(record['Điểm TBTL'])
                    if 0 <= score <= 4:
                        stats['scores'].append(score)
                except:
                    pass
        
        # Tính toán điểm
        if stats['scores']:
            stats['avg_score'] = sum(stats['scores']) / len(stats['scores'])
            stats['min_score'] = min(stats['scores'])
            stats['max_score'] = max(stats['scores'])
            stats['pass_rate'] = len([s for s in stats['scores'] if s >= 2.0]) / len(stats['scores']) * 100
        else:
            stats['avg_score'] = 0
            stats['min_score'] = 0
            stats['max_score'] = 0
            stats['pass_rate'] = 0
        
        return stats

def create_overview_metrics(stats):
    """Tạo metrics tổng quan."""
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("👥 Tổng số sinh viên", f"{stats['total_records']:,}")
    
    with col2:
        st.metric("📈 Điểm TB trung bình", f"{stats['avg_score']:.2f}")
    
    with col3:
        st.metric("✅ Tỷ lệ đạt (%)", f"{stats['pass_rate']:.1f}%")
    
    with col4:
        st.metric("📊 Số môn học", len(stats['by_subject']))

def main():
    st.markdown('<h1 style="text-align: center; color: #1f77b4;">📊 HỆ THỐNG QUẢN LÝ ĐIỂM - ĐHNN Huế</h1>', 
                unsafe_allow_html=True)
    st.markdown('<p style="text-align: center; color: #666;">Phiên Bản của Apus- hơi lỏ, thông cảm</p>', 
                unsafe_allow_html=True)
    
    processor = DataProcessor()
    
    # Load dữ liệu
    with st.spinner("Đang tải dữ liệu..."):
        data, error = processor.load_data_as_dict()
    
    if error:
        st.error(f"❌ {error}")
        st.info("💡 Chạy script direct_processor.py để tạo file output_direct.xlsx")
        return
    
    if not data:
        st.warning("⚠️ Không có dữ liệu")
        return
    
    stats = processor.analyze_data(data)
    
    tab1, tab2, tab3, tab4 = st.tabs(["📊 Tổng quan", "🔍 Tìm kiếm", "📋 Dữ liệu", "📤 Xuất file"])
    
    with tab1:
        st.subheader("📊 Thống kê tổng quan")
        create_overview_metrics(stats)
        
        st.markdown("---")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.subheader("📚 Thống kê theo khóa")
            for khoa, count in stats['by_khoa'].most_common():
                st.write(f"**{khoa}**: {count:,} sinh viên")
        
        with col2:
            st.subheader("📅 Thống kê theo học kỳ")
            for hk, count in stats['by_semester'].most_common():
                st.write(f"**{hk}**: {count:,} sinh viên")
        
        st.subheader("📚 Top 10 môn học")
        for i, (mon, count) in enumerate(stats['by_subject'].most_common(10), 1):
            st.write(f"{i}. **{mon}**: {count} sinh viên")
    
    with tab2:
        st.subheader("🔍 Tìm kiếm nâng cao")
        
        # Tìm kiếm chính
        col_main1, col_main2 = st.columns(2)
        
        with col_main1:
            main_search_name = st.text_input(
                "🏷️ Tìm theo họ tên:", 
                placeholder="Ví dụ: Thế Phú, Lê Thế",
                help="Tìm kiếm thông minh: có thể tìm từng từ riêng lẻ, không phân biệt hoa thường"
            )
        
        with col_main2:
            main_search_ma_sv = st.text_input(
                "🆔 Tìm theo mã sinh viên:", 
                placeholder="Ví dụ:  Mã qq gì đó",
                help="Hỗ trợ tìm kiếm một phần mã SV"
            )
        
        # Quick filters
        st.markdown("### ⚡ Bộ lọc nhanh")
        col_quick1, col_quick2, col_quick3, col_quick4 = st.columns(4)
        
        with col_quick1:
            quick_khoa = st.selectbox("Khóa:", ['Tất cả'] + sorted(list(stats['by_khoa'].keys())), key="quick_khoa")
        
        with col_quick2:
            quick_hk = st.selectbox("Học kỳ:", ['Tất cả'] + sorted(list(stats['by_semester'].keys())), key="quick_hk")
        
        with col_quick3:
            quick_status = st.selectbox("Trạng thái:", [
                'Tất cả', 'Đạt (≥ 2.0)', 'Không đạt (< 2.0)', 'Xuất sắc (≥ 3.6)'
            ], key="quick_status")
        
        with col_quick4:
            quick_mon = st.selectbox("Ngành:", ['Tất cả'] + sorted(list(stats['by_subject'].keys())[:20]), key="quick_mon")
        
        # Tìm kiếm
        search_results = data
        
        # Áp dụng tìm kiếm tên (thông minh)
        if main_search_name.strip():
            def smart_name_search(name_to_search, search_term):
                """Tìm kiếm thông minh: hỗ trợ tìm từng từ riêng lẻ"""
                name_lower = name_to_search.lower()
                search_lower = search_term.lower()
                
                # Tìm chính xác chuỗi con
                if search_lower in name_lower:
                    return True
                
                # Tìm từng từ riêng lẻ
                search_words = search_lower.split()
                name_words = name_lower.split()
                
                # Kiểm tra tất cả từ tìm kiếm có trong tên không
                for search_word in search_words:
                    found = False
                    for name_word in name_words:
                        if search_word in name_word or name_word in search_word:
                            found = True
                            break
                    if not found:
                        return False
                return True
            
            search_results = [r for r in search_results 
                            if smart_name_search(r.get('Họ và tên', ''), main_search_name)]
        
        # Áp dụng tìm kiếm mã SV
        if main_search_ma_sv.strip():
            search_results = [r for r in search_results 
                            if main_search_ma_sv.lower() in r.get('Mã SV', '').lower()]
        
        # Áp dụng quick filters
        if quick_khoa != 'Tất cả':
            search_results = [r for r in search_results if r.get('Khóa') == quick_khoa]
        
        if quick_hk != 'Tất cả':
            search_results = [r for r in search_results if r.get('Học kỳ') == quick_hk]
        
        if quick_mon != 'Tất cả':
            search_results = [r for r in search_results if r.get('Môn học') == quick_mon]
        
        if quick_status != 'Tất cả':
            filtered_by_status = []
            for r in search_results:
                try:
                    score = float(r.get('Điểm TBTL', 0))
                    if quick_status == 'Đạt (≥ 2.0)' and score >= 2.0:
                        filtered_by_status.append(r)
                    elif quick_status == 'Không đạt (< 2.0)' and score < 2.0:
                        filtered_by_status.append(r)
                    elif quick_status == 'Xuất sắc (≥ 3.6)' and score >= 3.6:
                        filtered_by_status.append(r)
                except:
                    pass
            search_results = filtered_by_status
        
        # Hiển thị kết quả
        st.markdown("---")
        
        if search_results:
            col_result1, col_result2 = st.columns([3, 1])
            with col_result1:
                st.success(f"🎯 Tìm thấy **{len(search_results):,}** kết quả phù hợp")
            with col_result2:
                show_all = st.checkbox("📋 Hiển thị tất cả", value=False, help="Hiển thị toàn bộ kết quả (có thể chậm nếu nhiều)")
            
            # Xác định số lượng kết quả hiển thị
            display_limit = len(search_results) if show_all else min(20, len(search_results))
            
            # Hiển thị chi tiết từng kết quả
            for i, record in enumerate(search_results[:display_limit]):
                with st.expander(f"#{i+1}: {record.get('Họ và tên', 'N/A')} - {record.get('Mã SV', 'N/A')}", expanded=False):
                    col_detail1, col_detail2 = st.columns(2)
                    
                    with col_detail1:
                        st.write("**👤 Thông tin sinh viên:**")
                        st.write(f"• **Họ tên:** {record.get('Họ và tên', 'N/A')}")
                        st.write(f"• **Mã SV:** {record.get('Mã SV', 'N/A')}")
                        st.write(f"• **Khóa:** {record.get('Khóa', 'N/A')}")
                        st.write(f"• **Học kỳ:** {record.get('Học kỳ', 'N/A')}")
                        st.write(f"• **Năm học:** {record.get('Năm học', 'N/A')}")
                    
                    with col_detail2:
                        st.write("**📊 Kết quả học tập:**")
                        st.write(f"• **Môn học:** {record.get('Môn học', 'N/A')}")
                        st.write(f"• **Điểm TBTL:** {record.get('Điểm TBTL', 'N/A')}")
                        st.write(f"• **Tổng TC:** {record.get('Tổng số tín chỉ', 'N/A')}")
                        st.write(f"• **TC lại:** {record.get('Số TC học/thi lại', 'N/A')}")
                        st.write(f"• **Xếp loại:** {record.get('Xếp loại học tập', 'N/A')}")
            
            # Thông báo trạng thái hiển thị
            if show_all:
                if len(search_results) > 20:
                    st.info(f"📋 Đang hiển thị tất cả **{len(search_results):,}** kết quả.")
            else:
                if len(search_results) > 20:
                    st.info(f"📝 Hiển thị **{display_limit}** / **{len(search_results):,}** kết quả. Tick ☑️ 'Hiển thị tất cả' để xem thêm.")
        else:
            st.warning("🔍 Không tìm thấy kết quả nào phù hợp với điều kiện tìm kiếm.")
            st.info("💡 Thử điều chỉnh từ khóa tìm kiếm hoặc bộ lọc.")
    
    with tab3:
        st.subheader("📋 Dữ liệu chi tiết")
        
        # Tìm kiếm
        st.markdown("### 🔍 Tìm kiếm")
        col_search1, col_search2 = st.columns(2)
        
        with col_search1:
            search_name = st.text_input("🏷️ Tìm theo tên sinh viên:", placeholder="Ví dụ: Thế Phú, Lê Thế", help="Tìm kiếm thông minh: có thể tìm từng từ riêng lẻ")
        
        with col_search2:
            search_ma_sv = st.text_input("🆔 Tìm theo mã sinh viên:", placeholder="Nhập mã sinh viên...")
        
        st.markdown("---")
        
        # Filters cơ bản
        st.markdown("### 📊 Lọc cơ bản")
        col1, col2, col3 = st.columns(3)
        
        with col1:
            khoa_options = ['Tất cả'] + sorted(list(stats['by_khoa'].keys()))
            selected_khoa = st.selectbox("Lọc theo khóa:", khoa_options)
        
        with col2:
            hk_options = ['Tất cả'] + sorted(list(stats['by_semester'].keys()))
            selected_hk = st.selectbox("Lọc theo học kỳ:", hk_options)
        
        with col3:
            mon_options = ['Tất cả'] + sorted(list(stats['by_subject'].keys())[:30])  # Top 30
            selected_mon = st.selectbox("Lọc theo môn:", mon_options)
        
        # Filter nâng cao
        with st.expander("🎛️ Bộ lọc nâng cao", expanded=False):
            col_adv1, col_adv2, col_adv3 = st.columns(3)
            
            with col_adv1:
                # Lọc theo điểm
                st.markdown("**📈 Lọc theo điểm TBTL:**")
                score_range = st.slider(
                    "Khoảng điểm:",
                    min_value=0.0,
                    max_value=4.0,
                    value=(0.0, 4.0),
                    step=0.1,
                    format="%.1f"
                )
                
                # Lọc theo tín chỉ
                st.markdown("**📚 Lọc theo tổng tín chỉ:**")
                # Tính min/max tín chỉ
                all_tc = []
                for record in data:
                    try:
                        tc = float(record.get('Tổng số tín chỉ', 0))
                        if tc > 0:
                            all_tc.append(tc)
                    except:
                        pass
                
                if all_tc:
                    min_tc, max_tc = int(min(all_tc)), int(max(all_tc))
                    tc_range = st.slider(
                        "Khoảng tín chỉ:",
                        min_value=min_tc,
                        max_value=max_tc,
                        value=(min_tc, max_tc),
                        step=1
                    )
                else:
                    tc_range = (0, 200)
            
            with col_adv2:
                # Lọc theo xếp loại
                st.markdown("**🏆 Lọc theo xếp loại học tập:**")
                xep_loai_options = set()
                for record in data:
                    xl = record.get('Xếp loại học tập', '')
                    if xl and xl.strip():
                        xep_loai_options.add(xl.strip())
                
                xep_loai_options = ['Tất cả'] + sorted(list(xep_loai_options))
                selected_xep_loai = st.selectbox("Xếp loại:", xep_loai_options)
                
                # Lọc theo trạng thái
                st.markdown("**📊 Lọc theo trạng thái:**")
                status_options = [
                    'Tất cả',
                    'Đạt (≥ 2.0)',
                    'Không đạt (< 2.0)',
                    'Xuất sắc (≥ 3.6)',
                    'Giỏi (3.2-3.59)',
                    'Khá (2.5-3.19)',
                    'Trung bình (2.0-2.49)'
                ]
                selected_status = st.selectbox("Trạng thái:", status_options)
            
            with col_adv3:
                # Lọc theo năm học
                st.markdown("**📅 Lọc theo năm học:**")
                nam_hoc_options = set()
                for record in data:
                    nh = record.get('Năm học', '')
                    if nh and nh.strip():
                        nam_hoc_options.add(nh.strip())
                
                if nam_hoc_options:
                    nam_hoc_options = ['Tất cả'] + sorted(list(nam_hoc_options))
                    selected_nam_hoc = st.selectbox("Năm học:", nam_hoc_options)
                else:
                    selected_nam_hoc = 'Tất cả'
                
                # Lọc theo số TC học/thi lại
                st.markdown("**🔄 Lọc theo TC học/thi lại:**")
                tc_lai_options = [
                    'Tất cả',
                    'Không có TC lại (= 0)',
                    'Có TC lại (> 0)',
                    'TC lại nhiều (≥ 10)'
                ]
                selected_tc_lai = st.selectbox("TC học/thi lại:", tc_lai_options)
        
        # Hàm hỗ trợ lọc
        def matches_search(record, search_name, search_ma_sv):
            """Kiểm tra xem record có match với tìm kiếm không (tìm kiếm thông minh)."""
            if search_name.strip():
                def smart_name_search(name_to_search, search_term):
                    name_lower = name_to_search.lower()
                    search_lower = search_term.lower()
                    
                    # Tìm chính xác chuỗi con
                    if search_lower in name_lower:
                        return True
                    
                    # Tìm từng từ riêng lẻ
                    search_words = search_lower.split()
                    name_words = name_lower.split()
                    
                    # Kiểm tra tất cả từ tìm kiếm có trong tên không
                    for search_word in search_words:
                        found = False
                        for name_word in name_words:
                            if search_word in name_word or name_word in search_word:
                                found = True
                                break
                        if not found:
                            return False
                    return True
                
                if not smart_name_search(record.get('Họ và tên', ''), search_name):
                    return False
            
            if search_ma_sv.strip():
                ma_sv = record.get('Mã SV', '').lower()
                if search_ma_sv.lower() not in ma_sv:
                    return False
            
            return True
        
        def matches_advanced_filters(record, score_range, tc_range, selected_xep_loai, 
                                   selected_status, selected_nam_hoc, selected_tc_lai):
            """Kiểm tra xem record có match với filter nâng cao không."""
            # Lọc theo điểm
            try:
                score = float(record.get('Điểm TBTL', 0))
                if not (score_range[0] <= score <= score_range[1]):
                    return False
            except:
                # Nếu không parse được điểm, bỏ qua filter điểm
                pass
            
            # Lọc theo tín chỉ
            try:
                tc = float(record.get('Tổng số tín chỉ', 0))
                if not (tc_range[0] <= tc <= tc_range[1]):
                    return False
            except:
                pass
            
            # Lọc theo xếp loại
            if selected_xep_loai != 'Tất cả':
                if record.get('Xếp loại học tập', '').strip() != selected_xep_loai:
                    return False
            
            # Lọc theo trạng thái điểm
            if selected_status != 'Tất cả':
                try:
                    score = float(record.get('Điểm TBTL', 0))
                    if selected_status == 'Đạt (≥ 2.0)' and score < 2.0:
                        return False
                    elif selected_status == 'Không đạt (< 2.0)' and score >= 2.0:
                        return False
                    elif selected_status == 'Xuất sắc (≥ 3.6)' and score < 3.6:
                        return False
                    elif selected_status == 'Giỏi (3.2-3.59)' and not (3.2 <= score < 3.6):
                        return False
                    elif selected_status == 'Khá (2.5-3.19)' and not (2.5 <= score < 3.2):
                        return False
                    elif selected_status == 'Trung bình (2.0-2.49)' and not (2.0 <= score < 2.5):
                        return False
                except:
                    pass
            
            # Lọc theo năm học
            if selected_nam_hoc != 'Tất cả':
                if record.get('Năm học', '').strip() != selected_nam_hoc:
                    return False
            
            # Lọc theo TC học/thi lại
            if selected_tc_lai != 'Tất cả':
                try:
                    tc_lai = float(record.get('Số TC học/thi lại', 0))
                    if selected_tc_lai == 'Không có TC lại (= 0)' and tc_lai != 0:
                        return False
                    elif selected_tc_lai == 'Có TC lại (> 0)' and tc_lai <= 0:
                        return False
                    elif selected_tc_lai == 'TC lại nhiều (≥ 10)' and tc_lai < 10:
                        return False
                except:
                    pass
            
            return True
        
        # Áp dụng tất cả các filter
        filtered_data = data
        
        # Lọc cơ bản
        if selected_khoa != 'Tất cả':
            filtered_data = [r for r in filtered_data if r.get('Khóa') == selected_khoa]
        
        if selected_hk != 'Tất cả':
            filtered_data = [r for r in filtered_data if r.get('Học kỳ') == selected_hk]
        
        if selected_mon != 'Tất cả':
            filtered_data = [r for r in filtered_data if r.get('Môn học') == selected_mon]
        
        # Lọc tìm kiếm
        if search_name.strip() or search_ma_sv.strip():
            filtered_data = [r for r in filtered_data if matches_search(r, search_name, search_ma_sv)]
        
        # Lọc nâng cao (chỉ áp dụng nếu expander được mở)
        filtered_data = [r for r in filtered_data if matches_advanced_filters(
            r, score_range, tc_range, selected_xep_loai, selected_status, selected_nam_hoc, selected_tc_lai
        )]
        
        # Tùy chọn hiển thị
        col_info, col_option = st.columns([3, 1])
        with col_info:
            st.info(f"Tìm thấy {len(filtered_data):,} / {len(data):,} bản ghi")
        with col_option:
            show_all_data = st.checkbox("📋 Hiển thị tất cả dữ liệu", value=False, help="Hiển thị toàn bộ dữ liệu (có thể chậm nếu nhiều)")
        
        # Xác định số lượng dữ liệu hiển thị
        data_limit = len(filtered_data) if show_all_data else min(100, len(filtered_data))
        display_data = filtered_data[:data_limit]
        
        if display_data:
            # Chuyển thành format cho st.table
            headers = list(display_data[0].keys())
            table_data = []
            for record in display_data:
                table_data.append([record.get(h, '') for h in headers])
            
            # Hiển thị bảng
            if show_all_data:
                st.write(f"**Dữ liệu đầy đủ ({len(display_data):,} dòng):**")
            else:
                st.write(f"**Dữ liệu mẫu ({len(display_data)} dòng đầu):**")
                
            import pandas as pd
            try:
                # Thử tạo DataFrame đơn giản
                df_display = pd.DataFrame(table_data, columns=headers)
                st.dataframe(df_display, use_container_width=True)
            except:
                # Fallback: hiển thị JSON
                st.write("**Dữ liệu (JSON format):**")
                st.json(display_data[:5])
            
            # Thông báo trạng thái
            if not show_all_data and len(filtered_data) > 100:
                st.info(f"📝 Hiển thị **{data_limit}** / **{len(filtered_data):,}** bản ghi. Tick ☑️ 'Hiển thị tất cả dữ liệu' để xem thêm.")
    
    with tab4:
        st.subheader("📤 Xuất dữ liệu")
        
        col1, col2 = st.columns(2)
        
        with col1:
            if st.button("💾 Xuất CSV", type="primary"):
                csv_path = processor.processing_path / "exported_data.csv"
                
                if data:
                    with open(str(csv_path), 'w', newline='', encoding='utf-8') as f:
                        if data:
                            writer = csv.DictWriter(f, fieldnames=data[0].keys())
                            writer.writeheader()
                            writer.writerows(data)
                    
                    st.success(f"✅ Đã xuất {len(data):,} bản ghi ra: {csv_path}")
        
        with col2:
            if st.button("📊 Xuất thống kê JSON"):
                json_path = processor.processing_path / "statistics.json"
                
                # Chuyển Counter thành dict để serialize
                export_stats = {
                    'total_records': stats['total_records'],
                    'avg_score': stats['avg_score'],
                    'pass_rate': stats['pass_rate'],
                    'by_khoa': dict(stats['by_khoa']),
                    'by_semester': dict(stats['by_semester']),
                    'by_subject': dict(stats['by_subject'])
                }
                
                with open(str(json_path), 'w', encoding='utf-8') as f:
                    json.dump(export_stats, f, ensure_ascii=False, indent=2)
                
                st.success(f"✅ Đã xuất thống kê ra: {json_path}")

if __name__ == "__main__":
    main()
